print("hello world")
import pandas as pd
import pandasql

# 默认读取第一个表单，直接默认读取到这个Excel的第一个表单
df = pd.read_excel('D:\\TAL\\考勤自动化\\2019年秋季教师请假汇总-更新2019.10.27-赵哲面授.xlsx', sheet_name='面授10.27')
df.columns = [
    'a1',
    'classid',
    'a3',
    'a4',
    'a5',
    'teacherid',
    'teacher',
    '学期',
    '年部',
    '年级',
    '年份',
    '学科',
    '班次',
    '服务中心',
    '教学点',
    '教室',
    '开课日期',
    '结课日期',
    '总课次',
    '课时',
    '上课时间',
    'kaoqin',
    '考勤次',
    '考勤原因',
    '代课教师姓名',
    '次数',
    '扣本次课时',
    '另扣课时',
    '考勤规则',
    '是否取消全勤',
    '是否取消晋升',
    'd1',
    'd2',
    '不另扣原因',
    '取消全勤',
    '取消晋升',
    '审核人'
]
data = df

sql = """
select  classid || teacherid,strftime('%Y-%m',kaoqin),count(*),sum(d1),sum(d2) from data where  a4='请假' and strftime('%Y-%m',kaoqin)="2019-10" group by  classid||teacherid,strftime('%Y-%m',kaoqin);
"""

dataframe = pandasql.sqldf(sql, globals())
# print(daframe)

import openpyxl

wb = openpyxl.load_workbook("D:\\TAL\\考勤自动化\\2019年10月秋季教师月考勤20191101-To薪资-示例.xlsx")
ws = wb["10月"]
rows = ws.max_row
columns = ws.max_column

for x in range(1, rows):
    # 获取表中x行1列的值
    class_id = ws.cell(row=x, column=2).value
    teacher_id = ws.cell(row=x, column=5).value
    uid = str(class_id) + str(teacher_id)
    # print(uid)
    for row in dataframe.values:
        # print(row[0], row[1], row[2], row[3])
        # 30~32
        if uid == row[0]:
            ws.cell(row=x, column=30).value = row[2]
            ws.cell(row=x, column=32).value = row[4]
            class_hours = ws.cell(row=x, column=15).value
            if class_hours * row[2] == row[3]:
                ws.cell(row=x, column=31).value = row[3]
            else:
                ws.cell(row=x, column=31).value = -100

wb.save("D:\\TAL\\考勤自动化\\2019年10月秋季教师月考勤20191101-To薪资-示例.xlsx")